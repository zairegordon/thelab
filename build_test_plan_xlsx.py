from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parent
OUTPUT = ROOT / "TEST_PLAN.xlsx"

navy = "17365D"
blue = "D9EAF7"
light_blue = "EAF3F8"
green = "E2F0D9"
yellow = "FFF2CC"
gray = "F2F2F2"
white = "FFFFFF"
red = "FCE4D6"
thin_gray = Side(style="thin", color="B7C9D6")

wb = Workbook()
cover = wb.active
cover.title = "Overview"
tests = wb.create_sheet("Test Cases")
data = wb.create_sheet("Test Data")
trace = wb.create_sheet("Traceability")
execution = wb.create_sheet("Execution Notes")
defects = wb.create_sheet("Defects")


def style_title(ws, title, subtitle=None):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    cell = ws.cell(1, 1, title)
    cell.font = Font(bold=True, size=16, color=white)
    cell.fill = PatternFill("solid", fgColor=navy)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
        sub = ws.cell(2, 1, subtitle)
        sub.font = Font(italic=True, color="44546A")
        sub.alignment = Alignment(wrap_text=True)


def style_headers(ws, row=1):
    for cell in ws[row]:
        cell.font = Font(bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin_gray)
    ws.row_dimensions[row].height = 30


def finish_sheet(ws, widths, freeze="A2", auto_filter=True):
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = freeze
    if auto_filter:
        ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin_gray)


# Overview
style_title(cover, "The Lab - Formal UI Test Plan", "Playwright browser test plan and execution sheet")
overview_rows = [
    ("Document type", "Formal browser UI test plan"),
    ("Application", "The Lab fantasy football comparison app"),
    ("Framework", "Playwright with Python and pytest"),
    ("Target browser", "Chromium"),
    ("Test environment", "Local Flask server on an ephemeral port"),
    ("Plan date", "2026-08-20"),
    ("Objective", "Verify the basic user-facing functionality currently implemented in The Lab."),
    ("Scope", "Homepage, player search, selection, navigation menus, data widgets, comparison setup, and responsive search layout."),
    ("Out of scope", "Third-party data accuracy, visual pixel comparison, calculation correctness, Sleeper upload processing, performance, and security testing."),
    ("Acceptance criteria", "UI-001 through UI-010 pass in Chromium and pytest exits with code 0."),
]
for index, (label, value) in enumerate(overview_rows, start=4):
    cover.cell(index, 1, label).font = Font(bold=True, color=navy)
    cover.cell(index, 2, value)
    cover.merge_cells(start_row=index, start_column=2, end_row=index, end_column=6)
finish_sheet(cover, {"A": 24, "B": 28, "C": 20, "D": 20, "E": 20, "F": 20}, freeze="A4", auto_filter=False)

# Test cases
style_title(tests, "Test Cases", "All Playwright scenarios covered by test_ui.py")
headers = ["Test ID", "Test Case", "Preconditions", "Test Steps", "Expected Result", "Status"]
for col, value in enumerate(headers, start=1):
    tests.cell(4, col, value)
style_headers(tests, 4)
case_rows = [
    ("UI-001", "Load homepage shell", "App dependencies installed", "Open `/`.", "Page title is The Lab; brand, player search field, and Search button are visible.", "Automated - pending clean run"),
    ("UI-002", "Render upcoming games ticker", "Homepage loaded; `/games` mocked with TD-006", "Open `/` and wait for ticker content.", "Ticker displays KC at BUF.", "Automated - pending clean run"),
    ("UI-003", "Search for a player", "Homepage loaded; `/suggest` mocked with TD-002", "Enter Mahomes in the player search field.", "A Patrick Mahomes suggestion appears.", "Automated - pending clean run"),
    ("UI-004", "Lock a player from suggestions", "UI-003 completed", "Click the Patrick Mahomes suggestion.", "URL contains the selected player identity and Selected players shows Patrick Mahomes.", "Automated - pending clean run"),
    ("UI-005", "Load Trending dropdown", "Homepage loaded; `/trending` mocked with TD-003", "Hover over Trending.", "Dropdown opens and displays Patrick Mahomes.", "Automated - pending clean run"),
    ("UI-006", "Load Injury Report dropdown", "Homepage loaded; `/injuries` mocked with TD-004", "Hover over Injury Report.", "Dropdown opens and displays an injury item for Patrick Mahomes.", "Automated - pending clean run"),
    ("UI-007", "Load Latest NFL News dropdown", "Homepage loaded; `/news` mocked with TD-005", "Hover over Latest NFL News.", "Dropdown opens and displays Mock NFL headline.", "Automated - pending clean run"),
    ("UI-008", "Open NFL Teams dropdown", "Homepage loaded", "Hover over NFL Teams.", "Team menu opens and contains a Kansas City Chiefs link for team=KC.", "Automated - pending clean run"),
    ("UI-009", "Expose comparison action", "Homepage loaded with TD-007 in the query string", "Open the page with both players selected.", "Selected players panel displays Compare trade action.", "Automated - pending clean run"),
    ("UI-010", "Preserve mobile search layout", "Homepage loaded at TD-008 viewport", "Set viewport to 390 x 844 and open `/`.", "Search surface remains within the viewport and search input remains visible.", "Automated - pending clean run"),
]
for row_index, row in enumerate(case_rows, start=5):
    for col, value in enumerate(row, start=1):
        tests.cell(row_index, col, value)
    tests.cell(row_index, 6).fill = PatternFill("solid", fgColor=yellow)
case_table = Table(displayName="TestCases", ref=f"A4:F{4 + len(case_rows)}")
case_table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
tests.add_table(case_table)
finish_sheet(tests, {"A": 12, "B": 30, "C": 38, "D": 46, "E": 52, "F": 28}, freeze="A5")

# Test data
style_title(data, "Test Data", "Deterministic fixtures used by the browser tests")
for col, value in enumerate(["Data ID", "Purpose", "Value"], start=1):
    data.cell(4, col, value)
style_headers(data, 4)
data_rows = [
    ("TD-001", "Search query", "Mahomes"),
    ("TD-002", "Suggested player", "Patrick Mahomes, QB, KC"),
    ("TD-003", "Trending player", "Patrick Mahomes, QB, KC, 42 adds"),
    ("TD-004", "Injury report", "Patrick Mahomes, Questionable, Limited practice"),
    ("TD-005", "News item", "Mock NFL headline"),
    ("TD-006", "Upcoming game", "2026, KC at BUF, Sun 1:00 PM"),
    ("TD-007", "Comparison players", "Christian McCaffrey and Patrick Mahomes"),
    ("TD-008", "Mobile viewport", "390 x 844 pixels"),
]
for row_index, row in enumerate(data_rows, start=5):
    for col, value in enumerate(row, start=1):
        data.cell(row_index, col, value)
data_table = Table(displayName="TestData", ref=f"A4:C{4 + len(data_rows)}")
data_table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
data.add_table(data_table)
finish_sheet(data, {"A": 14, "B": 24, "C": 64}, freeze="A5")

# Traceability
style_title(trace, "Traceability", "Requirement areas mapped to automated test cases")
for col, value in enumerate(["Requirement Area", "Covered By", "Coverage Description"], start=1):
    trace.cell(4, col, value)
style_headers(trace, 4)
trace_rows = [
    ("Homepage and search shell", "UI-001", "Brand, title, search input, and Search button"),
    ("Upcoming game ticker", "UI-002", "Mock games data renders in the ticker"),
    ("Player autocomplete", "UI-003", "Search suggestions appear from the mocked endpoint"),
    ("Player selection persistence", "UI-004", "Selected player identity is reflected in URL and panel"),
    ("Trending player data", "UI-005", "Trending dropdown opens and renders player data"),
    ("Injury report data", "UI-006", "Injury dropdown opens and renders report data"),
    ("NFL news data", "UI-007", "News dropdown opens and renders headline data"),
    ("NFL team navigation", "UI-008", "Teams dropdown contains expected team link"),
    ("Comparison setup", "UI-009", "Two selected players expose comparison action"),
    ("Responsive layout", "UI-010", "Search surface fits the mobile viewport"),
]
for row_index, row in enumerate(trace_rows, start=5):
    for col, value in enumerate(row, start=1):
        trace.cell(row_index, col, value)
trace_table = Table(displayName="Traceability", ref=f"A4:C{4 + len(trace_rows)}")
trace_table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
trace.add_table(trace_table)
finish_sheet(trace, {"A": 30, "B": 16, "C": 64}, freeze="A5")

# Execution notes
style_title(execution, "Execution Notes", "Environment, commands, and current validation state")
for col, value in enumerate(["Item", "Details", "Status"], start=1):
    execution.cell(4, col, value)
style_headers(execution, 4)
execution_rows = [
    ("Project dependency", "Playwright added to requirements.txt", "Complete"),
    ("Python package", "Playwright installed in project .venv", "Complete"),
    ("Browser runtime", "Chromium installed for Playwright", "Complete"),
    ("Test implementation", "test_ui.py contains UI-001 through UI-010", "Complete"),
    ("Clean pytest summary", "A complete final pytest summary was not captured because the terminal stopped returning reliable output after browser installation.", "Pending"),
    ("Backend test note", "Existing backend tests have unrelated failures in selected-player persistence, roster search ordering, and comparison scoring.", "Separate tracking"),
]
for row_index, row in enumerate(execution_rows, start=5):
    for col, value in enumerate(row, start=1):
        execution.cell(row_index, col, value)
    execution.cell(row_index, 3).fill = PatternFill("solid", fgColor=green if row[2] == "Complete" else yellow)
exec_table = Table(displayName="ExecutionNotes", ref=f"A4:C{4 + len(execution_rows)}")
exec_table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
execution.add_table(exec_table)
finish_sheet(execution, {"A": 24, "B": 100, "C": 22}, freeze="A5")

# Defects
style_title(defects, "Defect Log", "Use this sheet to record failures found during execution")
for col, value in enumerate(["Defect ID", "Related Test ID", "Summary", "Steps to Reproduce", "Expected Result", "Actual Result", "Severity", "Owner", "Resolution"], start=1):
    defects.cell(4, col, value)
style_headers(defects, 4)
defect_rows = [
    ("", "", "", "", "", "", "Critical / High / Medium / Low", "", ""),
]
for row_index, row in enumerate(defect_rows, start=5):
    for col, value in enumerate(row, start=1):
        defects.cell(row_index, col, value)
finish_sheet(defects, {"A": 14, "B": 18, "C": 30, "D": 42, "E": 34, "F": 34, "G": 20, "H": 18, "I": 30}, freeze="A5")

for ws in wb.worksheets:
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.outlinePr.summaryBelow = True

wb.save(OUTPUT)
load_workbook(OUTPUT).close()
print(f"Created {OUTPUT}")
